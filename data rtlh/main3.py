import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def adjust_column_width(worksheet):
    """
    Menyesuaikan lebar kolom berdasarkan konten
    """
    for column in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        adjusted_width = max_length + 2
        worksheet.column_dimensions[column_letter].width = adjusted_width

def clean_quotes(text):
    """
    Menghapus tanda petik dari teks
    """
    if isinstance(text, str):
        return text.replace('"', '')
    return text

def process_csv_to_excel_by_desa(folder_path):
    """
    Mengkonversi file CSV ke Excel dengan sheet terpisah untuk setiap desa
    """
    try:
        if not os.path.exists(folder_path):
            print(f"Folder tidak ditemukan: {folder_path}")
            return
        
        # Cari semua file CSV
        csv_files = [f for f in os.listdir(folder_path) if f.endswith('.csv')]
        
        if not csv_files:
            print("Tidak ada file CSV yang ditemukan dalam folder")
            return
            
        print(f"Ditemukan {len(csv_files)} file CSV")
        
        # Proses setiap file CSV (setiap kecamatan)
        for csv_file in csv_files:
            try:
                csv_path = os.path.join(folder_path, csv_file)
                excel_file = csv_file.replace('.csv', '.xlsx')
                excel_path = os.path.join(folder_path, excel_file)
                
                # Baca file CSV
                df = pd.read_csv(csv_path)
                
                # Hapus tanda petik dari seluruh kolom
                for column in df.columns:
                    df[column] = df[column].apply(clean_quotes)
                
                # Dapatkan nama kolom desa (kolom ke-5)
                desa_column = df.columns[4]  # index 4 = kolom ke-5
                
                # Buat Excel writer
                with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
                    # Grouped berdasarkan desa
                    for desa in df[desa_column].unique():
                        # Filter data untuk desa tersebut
                        desa_data = df[df[desa_column] == desa]
                        
                        # Tulis ke sheet baru
                        sheet_name = str(desa)[:31]  # Excel membatasi nama sheet maksimal 31 karakter
                        desa_data.to_excel(writer, sheet_name=sheet_name, index=False)
                        
                        # Sesuaikan lebar kolom
                        workbook = writer.book
                        worksheet = writer.sheets[sheet_name]
                        adjust_column_width(worksheet)
                
                print(f"Berhasil membuat file Excel: {excel_file}")
                print(f"- Dibuat {len(df[desa_column].unique())} sheet sesuai jumlah desa")
                
            except Exception as e:
                print(f"Error saat memproses {csv_file}: {str(e)}")
                continue
                
        print("\nProses konversi selesai!")
        
    except Exception as e:
        print(f"Error: {str(e)}")

# Contoh penggunaan
if __name__ == "__main__":
    folder_path = input("Masukkan path folder yang berisi file CSV: ")
    process_csv_to_excel_by_desa(folder_path)