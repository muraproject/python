import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def adjust_column_width(worksheet):
    """
    Menyesuaikan lebar kolom berdasarkan konten
    """
    for column in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        
        # Cek panjang setiap sel dalam kolom
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        # Sesuaikan lebar kolom dengan padding tambahan
        adjusted_width = max_length + 2
        worksheet.column_dimensions[column_letter].width = adjusted_width

def convert_csv_to_excel(folder_path):
    """
    Mengkonversi semua file CSV dalam folder ke Excel dengan lebar kolom otomatis
    """
    try:
        # Cek apakah folder ada
        if not os.path.exists(folder_path):
            print(f"Folder tidak ditemukan: {folder_path}")
            return
        
        # Cari semua file CSV
        csv_files = [f for f in os.listdir(folder_path) if f.endswith('.csv')]
        
        if not csv_files:
            print("Tidak ada file CSV yang ditemukan dalam folder")
            return
            
        print(f"Ditemukan {len(csv_files)} file CSV")
        
        # Proses setiap file CSV
        for csv_file in csv_files:
            try:
                csv_path = os.path.join(folder_path, csv_file)
                excel_file = csv_file.replace('.csv', '.xlsx')
                excel_path = os.path.join(folder_path, excel_file)
                
                # Baca file CSV
                df = pd.read_csv(csv_path)
                
                # Simpan ke Excel
                df.to_excel(excel_path, index=False)
                
                # Buka file Excel untuk menyesuaikan lebar kolom
                wb = load_workbook(excel_path)
                ws = wb.active
                
                # Sesuaikan lebar kolom
                adjust_column_width(ws)
                
                # Simpan perubahan
                wb.save(excel_path)
                
                print(f"Berhasil mengkonversi: {csv_file} -> {excel_file}")
                
            except Exception as e:
                print(f"Error saat memproses {csv_file}: {str(e)}")
                continue
                
        print("\nProses konversi selesai!")
        
    except Exception as e:
        print(f"Error: {str(e)}")

# Contoh penggunaan
if __name__ == "__main__":
    folder_path = input("Masukkan path folder yang berisi file CSV: ")
    convert_csv_to_excel(folder_path)