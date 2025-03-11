from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def create_university_excel():
    wb = Workbook()
    ws = wb.active
    
    # Set column widths
    for col in range(1, 8):
        ws.column_dimensions[get_column_letter(col)].width = 15
        
    # Create TABEL KODE JURUSAN
    ws['A1'] = 'TABEL KODE JURUSAN'
    ws.merge_cells('A1:D1')
    
    headers = ['KODE JURUSAN', 'JURUSAN', 'BAYAR PER SKS']
    for col, header in enumerate(headers, start=2):
        ws.cell(row=2, column=col, value=header)
    
    jurusan_data = [
        ('EK', 'EKONOMI', 13000),
        ('HK', 'HUKUM', 11000),
        ('SP', 'SOSIAL POLITIK', 12000),
        ('TA', 'TEKNIK ARSITEKTUR', 14000),
        ('TS', 'TEKNIK SIPIL', 15000)
    ]
    
    for row, data in enumerate(jurusan_data, start=3):
        for col, value in enumerate(data, start=2):
            ws.cell(row=row, column=col, value=value)
            
    # Create TABEL TAHUN ANGKATAN
    ws['A9'] = 'TABEL TAHUN ANGKATAN'
    ws.merge_cells('A9:C9')
    
    headers = ['TAHUN', 'ANGKATAN']
    for col, header in enumerate(headers, start=2):
        ws.cell(row=10, column=col, value=header)
    
    tahun_data = [
        ('87', '1987'),
        ('88', '1988'),
        ('89', '1989'),
        ('90', '1990')
    ]
    
    for row, data in enumerate(tahun_data, start=11):
        for col, value in enumerate(data, start=2):
            ws.cell(row=row, column=col, value=value)
            
    # Create TABEL JUMLAH SKS
    ws['E9'] = 'TABEL JUMLAH SKS'
    ws.merge_cells('E9:F9')
    
    headers = ['KODE SKS', 'JUMLAH SKS']
    for col, header in enumerate(headers, start=5):
        ws.cell(row=10, column=col, value=header)
    
    sks_data = [
        ('E20', 20),
        ('H18', 18),
        ('S21', 21),
        ('T20', 20)
    ]
    
    for row, data in enumerate(sks_data, start=11):
        for col, value in enumerate(data, start=5):
            ws.cell(row=row, column=col, value=value)
            
    # Create DAFTAR PEMBAYARAN UANG KULIAH
    ws['A16'] = 'DAFTAR PEMBAYARAN UANG KULIAH'
    ws.merge_cells('A16:G16')
    
    headers = ['NO', 'NO. POKOK', 'JURUSAN', 'ANGKATAN', 'BIAYA SKS', 'JUMLAH SKS', 'JUMLAH BIAYA']
    for col, header in enumerate(headers, start=1):
        ws.cell(row=17, column=col, value=header)
    
    no_pokok_data = [
        'TA89T20', 'TS90T20', 'HK87H18', 'EK88E20', 'SP87S21',
        'TS89T20', 'HK89H18', 'TA90T20', 'TA89T20', 'SP89S21'
    ]
    
    # Add NO and NO. POKOK data
    for row, no_pokok in enumerate(no_pokok_data, start=19):
        ws.cell(row=row, column=1, value=row-18)  # NO column
        ws.cell(row=row, column=2, value=no_pokok)  # NO. POKOK column
        
        # Add formulas
        # JURUSAN formula
        ws.cell(row=row, column=3, value=f'=VLOOKUP(LEFT(B{row},2),$B$3:$D$7,2)')
        
        # ANGKATAN formula
        ws.cell(row=row, column=4, value=f'=VLOOKUP(MID(B{row},3,2),$B$11:$C$14,2)')
        
        # BIAYA SKS formula
        ws.cell(row=row, column=5, value=f'=VLOOKUP(LEFT(B{row},2),$B$3:$D$7,3)')
        
        # JUMLAH SKS formula
        ws.cell(row=row, column=6, value=f'=VLOOKUP(RIGHT(B{row},3),$E$11:$F$14,2)')
        
        # JUMLAH BIAYA formula
        ws.cell(row=row, column=7, value=f'=E{row}*F{row}')
    
    # Add borders and styling
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin'))
    
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=7):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')
    
    # Save the workbook
    wb.save('Namaku.xlsx')

if __name__ == "__main__":
    create_university_excel()